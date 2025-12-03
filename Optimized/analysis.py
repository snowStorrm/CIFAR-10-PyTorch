import openpyxl

# Simply exports two arrays to an excel sheet for data analysis
# - Will append to an existing excel file
def exportResults(axisX: list[float], axisY: list[float], targetSheet: str=""):
    # Open the excel workbook
    try: wb = openpyxl.load_workbook('./out.xlsx')
    except: wb = openpyxl.Workbook()
    # Default excel sheet name
    if targetSheet == "": targetSheet = f"Sheet{len(wb.sheetnames)}"
    # Create a new sheet within the workbook if it doesn't exist already and open said sheet
    if targetSheet not in wb.sheetnames: wb.create_sheet(targetSheet)
    ws = wb[targetSheet]
    # Transpose the data and spacer arrays since we want data in columns but openpyxl works in rows
    a: list[str] = ["" for i in range(len(axisX))]
    for i in zip(axisX, axisY, a): ws.append(i)
    # Save and close
    wb.save('./out.xlsx')
    wb.close()



        